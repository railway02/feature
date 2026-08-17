#!/usr/bin/env python3
"""Freeze non-private api_fullseq_v3 Phase 1 artifacts and source provenance.

The restricted private label artifact is deliberately not opened, hashed, or
parsed.  Existing Phase 1 files are read-only; this script writes only the new
freeze manifest, freeze audit, and source-provenance table.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT = Path("/root/autodl-tmp/aneurysm")

PHASE1_ARTIFACTS = [
    ("code/25_build_lesion_series_registry.py", "phase1_builder"),
    ("metadata/api_fullseq_v3/lesion_registry_train_blinded.csv", "train_blinded_registry"),
    ("metadata/api_fullseq_v3/lesion_registry_valid_blinded.csv", "valid_blinded_registry"),
    ("reports/api_fullseq_v3/lesion_registry_audit.md", "registry_audit"),
    ("reports/api_fullseq_v3/lesion_series_manual_review.csv", "manual_review_seed"),
    ("reports/api_fullseq_v3/split_id_alignment_audit.md", "split_alignment_audit"),
]

EXPECTED = {
    "Train": {"rows": 1157, "patients": 1055},
    "Valid": {"rows": 289, "patients": 264},
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT)
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def uid_set_hash(values: pd.Series) -> tuple[int, str]:
    unique = sorted(set(values.astype(str)))
    return len(unique), sha256_text("\n".join(unique))


def canonical_cell(value: Any) -> dict[str, Any]:
    if value is None:
        return {"type": "null", "value": None}
    if isinstance(value, bool):
        return {"type": "bool", "value": value}
    if isinstance(value, datetime):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"type": "date", "value": value.isoformat()}
    if isinstance(value, float):
        if math.isnan(value):
            return {"type": "null", "value": None}
        if value.is_integer():
            return {"type": "int", "value": int(value)}
        return {"type": "float", "value": format(value, ".17g")}
    if isinstance(value, int):
        return {"type": "int", "value": value}
    return {"type": "str", "value": str(value)}


def normalize_patient_id(value: Any) -> str:
    if value is None:
        raise ValueError("Empty patient ID in source workbook")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if not text.isdigit():
        raise ValueError(f"Non-numeric patient ID: {value!r}")
    return str(int(text))


def atomic_write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.tmp")
    df.to_csv(
        temp,
        index=False,
        encoding="utf-8",
        lineterminator="\n",
        quoting=csv.QUOTE_MINIMAL,
    )
    os.replace(temp, path)


def atomic_write_text(text: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.tmp")
    temp.write_text(text, encoding="utf-8")
    os.replace(temp, path)


def inspect_artifact(path: Path, role: str, root: Path) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(path)
    suffix = path.suffix.casefold()
    row_count: int | str = ""
    column_count: int | str = ""
    text_line_count: int | str = ""
    lesion_uid_count: int | str = ""
    lesion_uid_hash = ""

    if suffix == ".csv":
        frame = pd.read_csv(path, dtype=str, keep_default_na=False)
        row_count = len(frame)
        column_count = len(frame.columns)
        if "lesion_uid" in frame.columns:
            lesion_uid_count, lesion_uid_hash = uid_set_hash(frame["lesion_uid"])
    elif suffix in {".md", ".py", ".txt", ".html"}:
        text_line_count = len(path.read_text(encoding="utf-8").splitlines())

    return {
        "relative_path": path.relative_to(root).as_posix(),
        "artifact_role": role,
        "artifact_format": suffix.lstrip("."),
        "sha256": sha256_file(path),
        "size_bytes": path.stat().st_size,
        "row_count": row_count,
        "column_count": column_count,
        "text_line_count": text_line_count,
        "lesion_uid_count": lesion_uid_count,
        "lesion_uid_set_sha256": lesion_uid_hash,
        "freeze_status": "frozen_read_only",
    }


def build_source_rows(
    root: Path,
    registry: pd.DataFrame,
    split: str,
    workbook_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook_hash = sha256_file(workbook_path)
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    if not workbook.sheetnames:
        raise ValueError(f"Workbook has no sheets: {workbook_path}")
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_material = json.dumps(
        [canonical_cell(value) for value in headers],
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    header_hash = sha256_text(header_material)

    provenance: list[dict[str, Any]] = []
    for row in registry.itertuples(index=False):
        physical_row = int(row.source_excel_row_id)
        if physical_row < 2 or physical_row > sheet.max_row:
            raise AssertionError(
                f"Registry row outside workbook bounds: {split} {row.lesion_uid} {physical_row}"
            )
        values = [
            sheet.cell(physical_row, column).value
            for column in range(1, sheet.max_column + 1)
        ]
        source_patient_id = normalize_patient_id(values[0])
        if source_patient_id != str(row.patient_id):
            raise AssertionError(
                f"Patient mismatch for {row.lesion_uid}: registry={row.patient_id}, source={source_patient_id}"
            )
        record_material = json.dumps(
            [
                {
                    "column_index": index,
                    "header": canonical_cell(header),
                    "cell": canonical_cell(value),
                }
                for index, (header, value) in enumerate(zip(headers, values), start=1)
            ],
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        provenance.append(
            {
                "lesion_uid": row.lesion_uid,
                "split": split,
                "patient_id": row.patient_id,
                "source_excel_relative_path": workbook_path.relative_to(root).as_posix(),
                "source_excel_sha256": workbook_hash,
                "source_sheet": sheet_name,
                "source_excel_row_id": physical_row,
                "source_header_sha256": header_hash,
                "source_record_content_sha256": sha256_text(record_material),
                "record_hash_canonicalization": "ordered_header_and_typed_cell_json_v1",
            }
        )
    workbook.close()
    summary = {
        "path": workbook_path.relative_to(root).as_posix(),
        "sha256": workbook_hash,
        "sheet": sheet_name,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "header_sha256": header_hash,
    }
    return provenance, summary


def main() -> int:
    args = parse_args()
    root = args.project_root.resolve()
    output_dir = root / "metadata/api_fullseq_v3"
    report_dir = root / "reports/api_fullseq_v3"
    manifest_path = output_dir / "phase1_frozen_manifest.csv"
    provenance_path = output_dir / "lesion_source_provenance.csv"
    audit_path = report_dir / "phase1_freeze_audit.md"

    private_path = output_dir / "lesion_outcome_labels_private.csv"
    private_exists = private_path.is_file()

    artifact_rows = [
        inspect_artifact(root / relative_path, role, root)
        for relative_path, role in PHASE1_ARTIFACTS
    ]
    artifact_frame = pd.DataFrame(artifact_rows)
    hashes_before = {
        row["relative_path"]: row["sha256"] for row in artifact_rows
    }

    train_path = output_dir / "lesion_registry_train_blinded.csv"
    valid_path = output_dir / "lesion_registry_valid_blinded.csv"
    train = pd.read_csv(train_path, dtype=str, keep_default_na=False)
    valid = pd.read_csv(valid_path, dtype=str, keep_default_na=False)
    for split, frame in (("Train", train), ("Valid", valid)):
        expected = EXPECTED[split]
        if len(frame) != expected["rows"] or frame["patient_id"].nunique() != expected["patients"]:
            raise AssertionError(f"Unexpected {split} registry scale")
        if not frame["lesion_uid"].is_unique:
            raise AssertionError(f"Duplicate lesion_uid in {split}")

    combined_uid_count, combined_uid_hash = uid_set_hash(
        pd.concat([train["lesion_uid"], valid["lesion_uid"]], ignore_index=True)
    )
    if combined_uid_count != 1446:
        raise AssertionError(f"Expected 1446 combined lesion UIDs, got {combined_uid_count}")

    train_provenance, train_source = build_source_rows(
        root, train, "Train", root / "metadata/Train.xlsx"
    )
    valid_provenance, valid_source = build_source_rows(
        root, valid, "Valid", root / "metadata/valid.xlsx"
    )
    provenance = pd.DataFrame(train_provenance + valid_provenance)
    if len(provenance) != 1446 or not provenance["lesion_uid"].is_unique:
        raise AssertionError("Source provenance scale/UID uniqueness failure")
    registry_uids = set(train["lesion_uid"]) | set(valid["lesion_uid"])
    if set(provenance["lesion_uid"]) != registry_uids:
        raise AssertionError("Source provenance lesion_uid set mismatch")

    audit_lines = [
        "# api_fullseq_v3 Phase 1 Freeze Audit",
        "",
        "## Freeze boundary",
        "",
        "- Existing Phase 1 builder, blinded registries, and reports were opened read-only and hashed.",
        "- The restricted private label artifact was not opened, parsed, or hashed in this phase.",
        f"- Restricted private artifact presence check only: {'present' if private_exists else 'absent'}.",
        "- New files are limited to the freeze manifest, source provenance, and this audit.",
        "",
        "## Frozen artifacts",
        "",
        "| Relative path | Role | SHA-256 | Rows | Columns | Text lines | lesion_uid count |",
        "|---|---|---|---:|---:|---:|---:|",
    ]
    for row in artifact_rows:
        audit_lines.append(
            f"| `{row['relative_path']}` | {row['artifact_role']} | `{row['sha256']}` | "
            f"{row['row_count'] or ''} | {row['column_count'] or ''} | "
            f"{row['text_line_count'] or ''} | {row['lesion_uid_count'] or ''} |"
        )
    audit_lines.extend(
        [
            "",
            "## Frozen lesion identity",
            "",
            f"- Combined lesion_uid count: **{combined_uid_count}**",
            f"- Combined lesion_uid set SHA-256: `{combined_uid_hash}`",
            f"- Train lesion_uid set SHA-256: `{uid_set_hash(train['lesion_uid'])[1]}`",
            f"- Valid lesion_uid set SHA-256: `{uid_set_hash(valid['lesion_uid'])[1]}`",
            "",
            "## Source provenance",
            "",
            f"- Train source: `{train_source['path']}`, sheet `{train_source['sheet']}`, "
            f"SHA-256 `{train_source['sha256']}`",
            f"- Valid source: `{valid_source['path']}`, sheet `{valid_source['sheet']}`, "
            f"SHA-256 `{valid_source['sha256']}`",
            f"- Provenance rows: **{len(provenance)}**",
            "- Each source-record hash covers the ordered header and typed cell values for the full physical row; raw source values are not copied into provenance.",
            "",
            "## Mutation check",
            "",
            "All frozen artifact hashes are recomputed after writing the new freeze outputs.",
            "",
        ]
    )

    atomic_write_csv(artifact_frame, manifest_path)
    atomic_write_csv(provenance, provenance_path)


    hashes_after = {
        relative_path: sha256_file(root / relative_path)
        for relative_path in hashes_before
    }
    if hashes_before != hashes_after:
        changed = sorted(
            path for path in hashes_before if hashes_before[path] != hashes_after[path]
        )
        raise AssertionError(f"Phase 1 artifacts changed during freeze: {changed}")
    audit_lines.extend(["- Post-write hash recomputation: **PASS**", "- Changed frozen artifacts: **0**", ""])
    atomic_write_text("\n".join(audit_lines), audit_path)

    print(
        json.dumps(
            {
                "frozen_artifacts": len(artifact_frame),
                "combined_lesion_uid_count": combined_uid_count,
                "combined_lesion_uid_set_sha256": combined_uid_hash,
                "source_provenance_rows": len(provenance),
                "restricted_private_artifact_content_accessed": False,
                "phase1_artifacts_changed": False,
                "outputs": [str(manifest_path), str(provenance_path), str(audit_path)],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
